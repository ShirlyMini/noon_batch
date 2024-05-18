import openpyxl

#### open xl-->sheet---> col and row
### create xl

# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# sh=wb['Sheet']
# sh.cell(1,1).value="name"
# sh.cell(2,1).value="pooja"
# sh.cell(3,1).value="sita"
# sh.cell(4,1).value="ram"
# wb.save("./test_xl2.xlsx")


####reading xl-- no

# wb = openpyxl.load_workbook("./test_xl.xlsx")
# print(wb.sheetnames)
# sh = wb['Sheet']
# print(sh.max_row)
# print(sh.max_column)
#
# for r in range(1, sh.max_row+1):
#     for c in range(1, sh.max_column+1):
#         print(sh.cell(r,c).value)

####write cell

wb = openpyxl.load_workbook("./test_xl.xlsx")
print(wb.sheetnames)
sh = wb['Sheet']

sh.cell(1,1).value="students"
sh.cell(5,1).value="nikil"
sh.cell(5,2).value="70"
sh.cell(6,1).value="sam"
sh.cell(6,2).value="75"

wb.save("./test_xl.xlsx")